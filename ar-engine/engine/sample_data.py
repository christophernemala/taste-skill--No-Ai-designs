"""Synthetic test data for the AR engine.

Generates the two input files in the exact layouts the engine expects,
deliberately reproducing all four known data traps:

  Trap 1 — intercompany accounts mixed into the extract,
  Trap 2 — system due dates equal to invoice dates (no terms in system),
  Trap 3 — terms that exist only in the prior-month workbook,
  Trap 4 — unapplied receipts / credit memos as negative open balances,
            plus near-duplicate customer names.

Entirely fictional customers and amounts. Safe to commit and to demo.
"""
import datetime as dt
import random

import openpyxl

import sys as _sys
from pathlib import Path as _P
_sys.path.insert(0, str(_P(__file__).resolve().parent.parent))
import config as cfg

random.seed(42)

EXTERNAL = [
    # (name, contract_terms or None, in_prior_master)
    ("Falcon Ridge Facilities Management LLC", 30, True),
    ("Oryx Marine Services W.L.L", 45, True),
    ("Al Sahra Retail Group LLC", 60, True),
    ("Al Sahra Retail Group WLL", None, False),          # near-duplicate of the above
    ("Dune Gate Contracting Co", 30, True),
    ("Pearl Quay Hospitality PJSC", 90, True),
    ("Coral Line Logistics FZE", 30, True),
    ("Ironwood Energy Solutions LLC", 45, True),
    ("Saffron Bay Trading EST", 30, True),
    ("Northlight Data Centres DMCC", 60, True),
    ("Mirage Valley Developments LLC", 30, True),
    ("Kestrel Aviation Support Services", None, False),  # new customer, no terms anywhere
    ("Bluewater Ports Authority", 120, True),
    ("Amber Dunes Healthcare Group", 30, True),
    ("Granite Peak Industrial LLC", 45, True),
]

INTERCOMPANY = [
    "MEA TECH GT LLC (Dubai Branch)",
    "Convergint Gulf Holdings (IC MEA 001)",
    "Convergint Qatar Operations (IC MEA 014)",
]

ENTITIES = ["DXB", "AUH", "IRQ"]


def _doc(i, ent):
    return f"INV/{ent}/{cfg.AS_OF.year}/{1000 + i}"


def make_extract():
    """NetSuite-style extract: headers row 7, data from row 8, customer on
    group-header rows only, 'Total - <customer>' rows closing each group."""
    cfg.DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARAgingDetailSOProjectID"
    ws.cell(1, 1, cfg.COMPANY_NAME)
    ws.cell(2, 1, "A/R Aging Detail SO/Project ID")
    ws.cell(3, 1, f"As of {cfg.ASOF_LABEL}")
    headers = ["Customer", "Transaction Type", "Date", "Document Number", "P.O. No.",
               "Due Date", "Age", "Open Balance", "Related Sales Order", "Project St"]
    for j, h in enumerate(headers, 1):
        ws.cell(7, j, h)

    r = 8
    doc_i = 0

    def group(name, docs):
        nonlocal r, doc_i
        ws.cell(r, 1, name)
        r += 1
        subtotal = 0.0
        for (ttype, inv_date, ns_due, amount) in docs:
            doc_i += 1
            ent = random.choice(ENTITIES)
            ws.cell(r, 2, ttype)
            ws.cell(r, 3, dt.datetime.combine(inv_date, dt.time()))
            ws.cell(r, 4, _doc(doc_i, ent) if ttype != "Payment" else f"PMT/{ent}/{2000 + doc_i}")
            ws.cell(r, 5, f"PO-{5000 + doc_i}")
            ws.cell(r, 6, dt.datetime.combine(ns_due, dt.time()))
            ws.cell(r, 7, max(0, (cfg.AS_OF - ns_due).days))
            ws.cell(r, 8, round(amount, 2))
            ws.cell(r, 9, f"SO-{3000 + doc_i}")
            ws.cell(r, 10, f"PRJ-{100 + doc_i % 40}")
            subtotal += amount
            r += 1
        ws.cell(r, 1, f"Total - {name}")
        ws.cell(r, 8, round(subtotal, 2))
        r += 1

    for name, terms, _ in EXTERNAL:
        docs = []
        n_inv = random.randint(3, 9)
        for _k in range(n_inv):
            days_back = random.randint(5, 320)
            inv_date = cfg.AS_OF - dt.timedelta(days=days_back)
            amount = random.choice([18, 35, 62, 90, 140, 260, 480]) * 1000 * random.uniform(0.7, 1.4)
            # Trap 2: most rows carry Due Date = Invoice Date (no terms in system)
            if random.random() < 0.72 or terms is None:
                ns_due = inv_date
            else:
                ns_due = inv_date + dt.timedelta(days=terms)
            docs.append(("Invoice", inv_date, ns_due, amount))
        # Trap 4: some customers carry unapplied receipts / credit memos
        if random.random() < 0.45:
            pay_date = cfg.AS_OF - dt.timedelta(days=random.randint(10, 90))
            docs.append((random.choice(["Payment", "Credit Memo"]), pay_date, pay_date,
                         -random.choice([15, 40, 75]) * 1000 * random.uniform(0.8, 1.3)))
        group(name, docs)

    # Trap 1: intercompany groups, large balances, journal-heavy
    for name in INTERCOMPANY:
        docs = []
        for _k in range(random.randint(4, 8)):
            days_back = random.randint(30, 400)
            inv_date = cfg.AS_OF - dt.timedelta(days=days_back)
            amount = random.choice([900, 1500, 2600, 4200]) * 1000 * random.uniform(0.8, 1.5)
            docs.append((random.choice(["Invoice", "Journal"]), inv_date, inv_date, amount))
        group(name, docs)

    wb.save(cfg.EXTRACT_FILE)
    return cfg.EXTRACT_FILE


def make_prior():
    """Prior-month workbook, SOA sheet: customer col 1, terms col 7, open
    balance col 10, data from row 8. This is the terms master (Trap 3)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg.PRIOR_SHEET
    ws.cell(1, 1, cfg.COMPANY_NAME)
    ws.cell(2, 1, f"Statement of Account - {cfg.PRIOR_LABEL}")
    for j, h in enumerate(["Customer", "", "", "", "", "", "Payment terms", "", "", "Open Balance"], 1):
        if h:
            ws.cell(7, j, h)
    r = 8
    for name, terms, in_prior in EXTERNAL:
        if not in_prior:
            continue
        bal = random.choice([120, 260, 400, 760, 1200]) * 1000 * random.uniform(0.8, 1.3)
        ws.cell(r, 1, name)
        ws.cell(r, 7, terms)
        ws.cell(r, 10, round(bal, 2))
        r += 1
    ws.cell(r, 1, "GRAND TOTAL")
    wb.save(cfg.PRIOR_FILE)
    return cfg.PRIOR_FILE


def run():
    e = make_extract()
    p = make_prior()
    print("Sample extract :", e)
    print("Sample prior   :", p)


if __name__ == "__main__":
    run()
