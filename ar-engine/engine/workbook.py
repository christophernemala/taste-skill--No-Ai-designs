"""Ten-tab formatted Excel AR pack, built from the engine JSON.

Tabs: Dashboard, Aging Summary, Customer Summary, Collection Forecast,
Invoice Detail (SOA), Terms Master, Exceptions & Mismatches,
Intercompany (Excluded), Methodology, RAW Data.

Formatting standard: dark navy headers, teal accent, centred cells,
money as #,##0.00 with red-parenthesised negatives, no gridlines,
freeze panes, autofilter, conditional formatting that earns its place.
"""
import datetime as dt
import json
from collections import Counter, defaultdict

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys as _sys
from pathlib import Path as _P
_sys.path.insert(0, str(_P(__file__).resolve().parent.parent))
import config as cfg

# ---------------- palette ----------------
NAVY = "0B1F3A"
NAVY2 = "13304F"
SLATE = "1C2C3E"
ACCENT = "00A9A5"
GOLD = "C9A227"
RED = "C0392B"
AMBER = "D68910"
GREEN = "1E8449"
LGREY = "F2F5F8"
WHITE = "FFFFFF"
BORDCOL = "B7C3D0"

thin = Side(style="thin", color=BORDCOL)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTRN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

MONEY = '#,##0.00;[Red](#,##0.00);"-"'
MONEY0 = '#,##0;[Red](#,##0);"-"'
PCT = "0.0%"
DATE = "DD-MMM-YYYY"

FC_KEYS = ["W1", "W2", "W3", "W4", "M1", "M2", "M3", "Beyond"]


def build(D=None):
    if D is None:
        D = json.load(open(cfg.ENGINE_JSON))
    ML = D["month_labels"]
    PRIOR = D["prior_label"]
    CUR = D["currency"]
    ASOF_L = D["asof_label"]
    BUCK = D["buckets"]
    INV = D["invoices"]
    TER = D["terms"]
    PRB = D["prior_balances"]
    PRT = D["prior_terms"]
    OUT = cfg.OUT_DIR / f"AR Aging Report - {ASOF_L}.xlsx"

    def bidx(b):
        return BUCK.index(b)

    def hdr(ws, row, headers, start=1, fill=NAVY, size=10, h=32):
        for i, t in enumerate(headers):
            c = ws.cell(row, start + i, t)
            c.font = Font(bold=True, color=WHITE, size=size, name="Calibri")
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = CTR
            c.border = BOX
        ws.row_dimensions[row].height = h

    def title(ws, row, text, sub=None, span=10):
        c = ws.cell(row, 1, text)
        c.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = CTRN
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        ws.row_dimensions[row].height = 30
        if sub:
            c = ws.cell(row + 1, 1, sub)
            c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
            c.fill = PatternFill("solid", fgColor=NAVY2)
            c.alignment = CTRN
            ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=span)
            ws.row_dimensions[row + 1].height = 20

    def widths(ws, spec):
        for col, w in spec.items():
            ws.column_dimensions[col].width = w

    def style_body(ws, r1, r2, c1, c2, money_cols=(), pct_cols=(), date_cols=(), banded=True):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                cell.alignment = CTRN
                cell.border = BOX
                cell.font = Font(size=10, name="Calibri")
                if c in money_cols:
                    cell.number_format = MONEY
                elif c in pct_cols:
                    cell.number_format = PCT
                elif c in date_cols:
                    cell.number_format = DATE
                if banded and (r - r1) % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=LGREY)

    def totrow(ws, row, c1, c2, label_col=1, label="GRAND TOTAL"):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row, c)
            cell.font = Font(bold=True, size=10.5, color=WHITE, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=SLATE)
            cell.alignment = CTRN
            cell.border = Border(left=thin, right=thin, top=Side(style="double", color=NAVY), bottom=thin)
        ws.cell(row, label_col).value = label

    # ---- customer roll-up ----
    cust = defaultdict(lambda: dict(b=[0.0] * 6, open=0.0, n=0, oldest=0,
                                    fc={k: 0.0 for k in FC_KEYS}, neg=0.0, negn=0))
    for r in INV:
        c = cust[r["customer"]]
        c["b"][bidx(r["bucket"])] += r["open_bal"]
        c["open"] += r["open_bal"]
        c["n"] += 1
        c["oldest"] = max(c["oldest"], r["age"])
        c["fc"][r["fc_month"]] += r["open_bal"]
        if r["fc_month"] == "M1" and r["fc_week"]:
            c["fc"]["W%d" % r["fc_week"]] += r["open_bal"]
        if r["open_bal"] < 0:
            c["neg"] += r["open_bal"]
            c["negn"] += 1

    names = sorted(cust, key=lambda x: -cust[x]["open"])
    TOTAL = sum(cust[c]["open"] for c in cust)
    TOTB = [sum(cust[c]["b"][i] for c in cust) for i in range(6)]
    PRIOR_TOTAL = sum(PRB.values())
    ic_total = sum(x["open_bal"] for x in D["ic"])
    ic_accounts = len({x["customer"] for x in D["ic"]})

    def risk(c):
        d = cust[c]
        if d["open"] <= 0:
            return "Credit Balance"
        p = (d["b"][3] + d["b"][4] + d["b"][5]) / d["open"] if d["open"] else 0
        if p >= 0.75 or d["oldest"] > 365:
            return "Critical"
        if p >= 0.40 or d["oldest"] > 180:
            return "High"
        if p >= 0.15 or d["oldest"] > 90:
            return "Medium"
        return "Low"

    def prio(c):
        d = cust[c]
        v = d["b"][3] + d["b"][4] + d["b"][5]
        if v >= cfg.PRIO_P1:
            return "P1 - Escalate"
        if v >= cfg.PRIO_P2:
            return "P2 - Call this week"
        if v > 0:
            return "P3 - Email reminder"
        return "P4 - Monitor"

    wb = openpyxl.Workbook()

    # =====================================================================
    # 1. DASHBOARD
    # =====================================================================
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    for r in range(1, 70):
        for c in range(1, 20):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=NAVY)
    widths(ws, {get_column_letter(i): 16 for i in range(1, 20)})
    ws.column_dimensions["A"].width = 3

    def dcell(r, c, v, size=11, bold=False, color=WHITE, fmt=None, fill=None, merge=None, align="center"):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, size=size, color=color, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fmt:
            cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=fill or NAVY)
        if merge:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge)
        return cell

    dcell(2, 2, D["company"].upper(), 20, True, WHITE, merge=17)
    dcell(3, 2, "ACCOUNTS RECEIVABLE AGING & COLLECTION DASHBOARD", 13, True, ACCENT, merge=17)
    dcell(4, 2, f"As of {ASOF_L}   |   Currency: {CUR}   |   Intercompany balances excluded   |   Prepared by {cfg.PREPARED_BY}",
          9.5, False, "A9BFD4", merge=17)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18

    past = TOTAL - TOTB[0]
    o90 = TOTB[4] + TOTB[5]
    o60 = TOTB[3] + o90
    neg = sum(r["open_bal"] for r in INV if r["open_bal"] < 0)
    negn = sum(1 for r in INV if r["open_bal"] < 0)

    kpis = [
        ("TOTAL OPEN AR", TOTAL, MONEY0, ACCENT, f"{len(names)} customers / {len(INV):,} documents"),
        ("CURRENT (NOT DUE)", TOTB[0], MONEY0, GREEN, f"{TOTB[0]/TOTAL:.1%} of book" if TOTAL else "-"),
        ("TOTAL PAST DUE", past, MONEY0, AMBER, f"{past/TOTAL:.1%} of book" if TOTAL else "-"),
        ("OVER 90 DAYS", o90, MONEY0, RED, f"{o90/TOTAL:.1%} - provision review" if TOTAL else "-"),
        ("UNAPPLIED CR/CASH", neg, MONEY0, GOLD, f"{negn} documents to allocate"),
        ("MoM MOVEMENT", TOTAL - PRIOR_TOTAL, MONEY0, "5DADE2", f"vs {PRIOR} {PRIOR_TOTAL:,.0f}"),
    ]
    row = 6
    for i, (lab, val, fmt, col, note) in enumerate(kpis):
        c0 = 2 + i * 3
        dcell(row, c0, lab, 9, True, "A9BFD4", fill=NAVY2, merge=c0 + 2)
        dcell(row + 1, c0, val, 17, True, col, fmt=fmt, fill=NAVY2, merge=c0 + 2)
        dcell(row + 2, c0, note, 8.5, False, "A9BFD4", fill=NAVY2, merge=c0 + 2)
        for rr in range(row, row + 3):
            for cc in range(c0, c0 + 3):
                ws.cell(rr, cc).border = Border(left=Side("thin", color=ACCENT), right=Side("thin", color=ACCENT),
                                                top=Side("thin", color=ACCENT), bottom=Side("thin", color=ACCENT))
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 16

    # aging table + charts
    r0 = 11
    dcell(r0, 2, "AGING PROFILE", 12, True, ACCENT, merge=8)
    hdr(ws, r0 + 1, ["Bucket", f"Amount ({CUR})", "% of Total"], start=2, fill=NAVY2, size=9.5, h=26)
    for i, b in enumerate(BUCK):
        rr = r0 + 2 + i
        dcell(rr, 2, b, 10, True, WHITE, fill=SLATE)
        dcell(rr, 3, TOTB[i], 10, False, WHITE, fmt=MONEY0, fill=SLATE)
        dcell(rr, 4, TOTB[i] / TOTAL if TOTAL else 0, 10, False, WHITE, fmt=PCT, fill=SLATE)
    rr = r0 + 8
    dcell(rr, 2, "TOTAL", 10.5, True, ACCENT, fill=NAVY2)
    dcell(rr, 3, TOTAL, 10.5, True, ACCENT, fmt=MONEY0, fill=NAVY2)
    dcell(rr, 4, 1, 10.5, True, ACCENT, fmt=PCT, fill=NAVY2)

    ch = BarChart()
    ch.type = "col"
    ch.title = f"AR by Aging Bucket ({CUR})"
    ch.add_data(Reference(ws, min_col=3, min_row=r0 + 1, max_row=r0 + 7), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=r0 + 2, max_row=r0 + 7))
    ch.height, ch.width = 8.2, 17
    ch.legend = None
    ws.add_chart(ch, "H12")

    pie = PieChart()
    pie.title = "Aging Mix"
    pie.add_data(Reference(ws, min_col=3, min_row=r0 + 1, max_row=r0 + 7), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=2, min_row=r0 + 2, max_row=r0 + 7))
    pie.height, pie.width = 8.2, 11
    ws.add_chart(pie, "N12")

    # top 15
    r1 = 22
    dcell(r1, 2, "TOP 15 CUSTOMERS BY OPEN AR", 12, True, ACCENT, merge=10)
    hdr(ws, r1 + 1, ["#", "Customer", "Open AR", "Current", "Past Due", "Over 90", "Overdue %", "Risk", "Priority"],
        start=2, fill=NAVY2, size=9.5, h=26)
    for i, c in enumerate(names[:15]):
        d = cust[c]
        rr = r1 + 2 + i
        pd_ = d["open"] - d["b"][0]
        dcell(rr, 2, i + 1, 9.5, False, WHITE, fill=SLATE)
        dcell(rr, 3, c[:52], 9.5, False, WHITE, fill=SLATE, align="left")
        dcell(rr, 4, d["open"], 9.5, True, WHITE, fmt=MONEY0, fill=SLATE)
        dcell(rr, 5, d["b"][0], 9.5, False, WHITE, fmt=MONEY0, fill=SLATE)
        dcell(rr, 6, pd_, 9.5, False, AMBER, fmt=MONEY0, fill=SLATE)
        dcell(rr, 7, d["b"][4] + d["b"][5], 9.5, False, RED, fmt=MONEY0, fill=SLATE)
        dcell(rr, 8, pd_ / d["open"] if d["open"] else 0, 9.5, False, WHITE, fmt=PCT, fill=SLATE)
        dcell(rr, 9, risk(c), 9.5, False, WHITE, fill=SLATE)
        dcell(rr, 10, prio(c), 9.5, False, WHITE, fill=SLATE)
    ws.column_dimensions["C"].width = 46

    ch2 = BarChart()
    ch2.type = "bar"
    ch2.title = f"Top 15 Customers - Open AR ({CUR})"
    ch2.add_data(Reference(ws, min_col=4, min_row=r1 + 1, max_row=r1 + 16), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=3, min_row=r1 + 2, max_row=r1 + 16))
    ch2.height, ch2.width = 12, 15
    ch2.legend = None
    ws.add_chart(ch2, "M24")

    # forecast strip
    r2 = 40
    FC = {k: sum(cust[c]["fc"][k] for c in cust) for k in FC_KEYS}
    fc_total = FC["M1"] + FC["M2"] + FC["M3"] + FC["Beyond"]
    dcell(r2, 2, "COLLECTION FORECAST", 12, True, ACCENT, merge=10)
    hdr(ws, r2 + 1, [f"{ML['M1']} W1", f"{ML['M1']} W2", f"{ML['M1']} W3", f"{ML['M1']} W4",
                     f"{ML['M1'].upper()} TOTAL", ML["M2"], ML["M3"], ML["Beyond"], "TOTAL"],
        start=2, fill=NAVY2, size=9.5, h=26)
    vals = [FC["W1"], FC["W2"], FC["W3"], FC["W4"], FC["M1"], FC["M2"], FC["M3"], FC["Beyond"], fc_total]
    for i, v in enumerate(vals):
        dcell(r2 + 2, 2 + i, v, 11, True, ACCENT if i in (4, 8) else WHITE, fmt=MONEY0,
              fill=NAVY2 if i in (4, 8) else SLATE)
    dcell(r2 + 3, 2, "Forecast reconciles to Open AR:", 9, True, "A9BFD4", merge=5, align="right")
    dcell(r2 + 3, 7, fc_total - TOTAL, 9, True, GREEN, fmt=MONEY, merge=9)

    # data integrity
    dcell(46, 2, "DATA INTEGRITY", 12, True, ACCENT, merge=10)
    checks = [
        ("Aging buckets reconcile to Open AR", abs(sum(TOTB) - TOTAL) < 0.01),
        ("Forecast reconciles to Open AR", abs(fc_total - TOTAL) < 0.01),
        (f"Intercompany excluded ({ic_accounts} accounts, {CUR} {ic_total:,.0f})", True),
        (f"Payment terms confirmed from {PRIOR} master: "
         f"{sum(1 for v in TER.values() if v['status'] == 'Confirmed')} of {len(TER)} customers", True),
        (f"Terms assumed / derived (see Terms Master tab): "
         f"{sum(1 for v in TER.values() if v['status'] != 'Confirmed')}", False),
        (f"Unapplied cash / credit memos to allocate: {negn} docs, {CUR} {neg:,.0f}", False),
        (f"Near-duplicate customer names flagged: {len(D['near_pairs']) + len(D['dup_groups'])}", False),
    ]
    for i, (t, ok) in enumerate(checks):
        dcell(47 + i, 2, ("PASS   " if ok else "REVIEW ") + t, 10, True, GREEN if ok else GOLD,
              fill=SLATE, merge=11, align="left")
    ws.freeze_panes = "A6"

    # =====================================================================
    # 2. AGING SUMMARY
    # =====================================================================
    ws = wb.create_sheet("Aging Summary")
    ws.sheet_view.showGridLines = False
    title(ws, 1, D["company"].upper(),
          f"Aged Receivables Summary as of {ASOF_L}  |  {CUR}  |  Intercompany excluded", span=9)
    hdr(ws, 4, ["Aging Bucket", "Days Past Due", f"Open Balance ({CUR})", "% of Total AR", "No. of Documents",
                "No. of Customers", "Cumulative %", "Collectability", "Action Required"])
    ranges = ["Not yet due", "1 - 30 days", "31 - 60 days", "61 - 90 days", "91 - 120 days", "Over 120 days"]
    collect = ["Expected on terms", "High", "High", "Moderate", "Low - review", "Doubtful - provision"]
    action = ["Monitor", "Statement / reminder", "Collector call", "Escalate to PM & Sales",
              "Escalate to Finance Manager", "Legal / provision assessment"]
    docn = Counter(r["bucket"] for r in INV)
    cusn = {b: len({r["customer"] for r in INV if r["bucket"] == b}) for b in BUCK}
    cum = 0
    for i, b in enumerate(BUCK):
        r = 5 + i
        cum += TOTB[i] / TOTAL if TOTAL else 0
        for j, v in enumerate([b, ranges[i], TOTB[i], TOTB[i] / TOTAL if TOTAL else 0,
                               docn[b], cusn[b], cum, collect[i], action[i]], 1):
            ws.cell(r, j, v)
    style_body(ws, 5, 10, 1, 9, money_cols=(3,), pct_cols=(4, 7))
    ws.cell(11, 3, TOTAL)
    ws.cell(11, 4, 1)
    ws.cell(11, 5, len(INV))
    ws.cell(11, 6, len(names))
    ws.cell(11, 7, 1)
    totrow(ws, 11, 1, 9, label="TOTAL OPEN AR")
    ws.cell(11, 3).number_format = MONEY
    ws.cell(11, 4).number_format = PCT
    ws.cell(11, 7).number_format = PCT

    hdr(ws, 14, ["Key Metric", ASOF_L, PRIOR, "Movement", "Comment"])
    mets = [
        (f"Total Open AR ({CUR})", TOTAL, PRIOR_TOTAL, TOTAL - PRIOR_TOTAL, "External customers only"),
        ("Current / not yet due", TOTB[0], None, None, f"{TOTB[0]/TOTAL:.1%} of book" if TOTAL else "-"),
        ("Total past due", past, None, None, f"{past/TOTAL:.1%} of book" if TOTAL else "-"),
        ("Over 60 days", o60, None, None, f"{o60/TOTAL:.1%} of book" if TOTAL else "-"),
        ("Over 90 days", o90, None, None, "Provision review population"),
        ("Over 120 days", TOTB[5], None, None, "Legal / write-off assessment"),
        ("Unapplied cash & credit memos", neg, None, None, f"{negn} documents pending allocation"),
        ("Active customers", len(names), len(PRB), len(names) - len(PRB), "Accounts with an open balance"),
        ("Open documents", len(INV), None, None, "Invoices, journals, payments, credit memos"),
    ]
    r = 15
    for m in mets:
        for j, v in enumerate(m, 1):
            ws.cell(r, j, v)
        r += 1
    style_body(ws, 15, r - 1, 1, 5, money_cols=(2, 3, 4))
    for rr in (22, 23):
        for c in (2, 3, 4):
            ws.cell(rr, c).number_format = "#,##0"
    widths(ws, {"A": 30, "B": 20, "C": 20, "D": 18, "E": 42, "F": 16, "G": 14, "H": 24, "I": 32})
    ws.conditional_formatting.add("C5:C10", ColorScaleRule(start_type="min", start_color="E8F6F3",
                                                           end_type="max", end_color="F5B7B1"))

    # =====================================================================
    # 3. CUSTOMER SUMMARY
    # =====================================================================
    ws = wb.create_sheet("Customer Summary")
    ws.sheet_view.showGridLines = False
    H = ["S.No", "Customer", "Payment Terms (Days)", "Terms Source", "Open AR", "Current", "0-30", "31-60",
         "61-90", "91-120", "Over 120", "Bucket Check", "Total Past Due", "Overdue %", "Over 60", "Over 90",
         "Over 120 %", "Oldest Age (Days)", "Documents", "Credit / Unapplied", "Risk Category",
         "Collection Priority", "Recommended Action", f"{PRIOR} Balance", "MoM Movement", "Remarks"]
    title(ws, 1, "CUSTOMER AGING SUMMARY", f"One row per customer  |  As of {ASOF_L}  |  {CUR}", span=len(H))
    hdr(ws, 4, H)
    ACT = {"P1 - Escalate": "Escalate to CFO / legal review",
           "P2 - Call this week": "Collector call + statement of account",
           "P3 - Email reminder": "Send reminder with SOA",
           "P4 - Monitor": "Monitor - no action required"}
    r = 5
    for i, c in enumerate(names, 1):
        d = cust[c]
        t = TER[c]
        pd_ = d["open"] - d["b"][0]
        remarks = []
        if t["status"] != "Confirmed":
            remarks.append(t["status"])
        if d["negn"]:
            remarks.append(f"{d['negn']} credit/payment doc(s) unapplied")
        if c not in PRB:
            remarks.append(f"New customer vs {PRIOR}")
        vals = [i, c, t["terms"], t["source"], d["open"], *d["b"],
                round(sum(d["b"]) - d["open"], 2), pd_, (pd_ / d["open"] if d["open"] else 0),
                d["b"][3] + d["b"][4] + d["b"][5], d["b"][4] + d["b"][5],
                (d["b"][5] / d["open"] if d["open"] else 0), d["oldest"], d["n"], d["neg"],
                risk(c), prio(c), ACT[prio(c)], PRB.get(c),
                (d["open"] - PRB[c]) if c in PRB else None,
                "; ".join(remarks)]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v)
        r += 1
    last = r - 1
    style_body(ws, 5, last, 1, len(H), money_cols=(5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 20, 24, 25),
               pct_cols=(14, 17))
    for rr in range(5, last + 1):
        for cc in (2, 4, 23, 26):
            ws.cell(rr, cc).alignment = LEFT
    tr = last + 1
    ws.cell(tr, 5, TOTAL)
    for i in range(6):
        ws.cell(tr, 6 + i, TOTB[i])
    ws.cell(tr, 12, round(sum(TOTB) - TOTAL, 2))
    ws.cell(tr, 13, past)
    ws.cell(tr, 14, past / TOTAL if TOTAL else 0)
    ws.cell(tr, 15, o60)
    ws.cell(tr, 16, o90)
    ws.cell(tr, 17, TOTB[5] / TOTAL if TOTAL else 0)
    ws.cell(tr, 19, len(INV))
    ws.cell(tr, 20, neg)
    ws.cell(tr, 24, PRIOR_TOTAL)
    ws.cell(tr, 25, TOTAL - PRIOR_TOTAL)
    totrow(ws, tr, 1, len(H))
    for c in (5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 20, 24, 25):
        ws.cell(tr, c).number_format = MONEY
    for c in (14, 17):
        ws.cell(tr, c).number_format = PCT
    widths(ws, {"A": 6, "B": 48, "C": 11, "D": 34, "E": 16, "F": 15, "G": 14, "H": 14, "I": 14, "J": 14,
                "K": 15, "L": 12, "M": 16, "N": 11, "O": 15, "P": 15, "Q": 11, "R": 13, "S": 11, "T": 16,
                "U": 14, "V": 20, "W": 34, "X": 16, "Y": 15, "Z": 40})
    ws.auto_filter.ref = f"A4:{get_column_letter(len(H))}{last}"
    ws.freeze_panes = "C5"
    ws.conditional_formatting.add(f"U5:U{last}", CellIsRule(operator="equal", formula=['"Critical"'],
        fill=PatternFill("solid", fgColor="F5B7B1"), font=Font(bold=True, color="922B21")))
    ws.conditional_formatting.add(f"U5:U{last}", CellIsRule(operator="equal", formula=['"High"'],
        fill=PatternFill("solid", fgColor="FAD7A0"), font=Font(bold=True, color="9C640C")))
    ws.conditional_formatting.add(f"U5:U{last}", CellIsRule(operator="equal", formula=['"Low"'],
        fill=PatternFill("solid", fgColor="D5F5E3"), font=Font(bold=True, color="196F3D")))
    ws.conditional_formatting.add(f"L5:L{last}", CellIsRule(operator="notEqual", formula=["0"],
        fill=PatternFill("solid", fgColor="F5B7B1"), font=Font(bold=True, color="922B21")))
    ws.conditional_formatting.add(f"E5:E{last}", ColorScaleRule(start_type="min", start_color="FFFFFF",
                                                                end_type="max", end_color="7FB3D5"))

    # =====================================================================
    # 4. COLLECTION FORECAST
    # =====================================================================
    ws = wb.create_sheet("Collection Forecast")
    ws.sheet_view.showGridLines = False
    H = ["S.No", "Customer", "Terms", "Open AR", "Current", "0-30", "31-60", "61-90", "91-120", "Over 120",
         f"{ML['M1']} W1", f"{ML['M1']} W2", f"{ML['M1']} W3", f"{ML['M1']} W4", f"Total {ML['M1']}",
         ML["M2"], ML["M3"], ML["Beyond"], "Forecast Total", "Variance to Open AR",
         "Status", "Collector Override", "Collector Comment"]
    title(ws, 1, "COLLECTION FORECAST",
          "Weekly and monthly expected receipts  |  Baseline is system-generated - collector may override in the last two columns",
          span=len(H))
    hdr(ws, 4, H)
    r = 5
    for i, c in enumerate(names, 1):
        d = cust[c]
        f = d["fc"]
        ftot = f["M1"] + f["M2"] + f["M3"] + f["Beyond"]
        vals = [i, c, TER[c]["terms"], d["open"], *d["b"],
                f["W1"], f["W2"], f["W3"], f["W4"], f["M1"], f["M2"], f["M3"], f["Beyond"],
                ftot, round(d["open"] - ftot, 2),
                "Fully Allocated" if abs(d["open"] - ftot) < 0.01 else "Review Required", None, None]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v)
        r += 1
    last = r - 1
    style_body(ws, 5, last, 1, len(H), money_cols=tuple(range(4, 21)) + (22,))
    for rr in range(5, last + 1):
        ws.cell(rr, 2).alignment = LEFT
    tr = last + 1
    ws.cell(tr, 4, TOTAL)
    for i in range(6):
        ws.cell(tr, 5 + i, TOTB[i])
    for j, k in enumerate(FC_KEYS):
        ws.cell(tr, 11 + j, FC[k])
    ws.cell(tr, 19, fc_total)
    ws.cell(tr, 20, round(TOTAL - fc_total, 2))
    ws.cell(tr, 21, "Fully Allocated")
    totrow(ws, tr, 1, len(H))
    for c in range(4, 21):
        ws.cell(tr, c).number_format = MONEY
    widths(ws, {"A": 6, "B": 48, "C": 8, "D": 16, "E": 15, "F": 14, "G": 14, "H": 14, "I": 14, "J": 15,
                "K": 14, "L": 14, "M": 14, "N": 14, "O": 16, "P": 15, "Q": 15, "R": 16, "S": 16, "T": 16,
                "U": 17, "V": 17, "W": 34})
    ws.auto_filter.ref = f"A4:{get_column_letter(len(H))}{last}"
    ws.freeze_panes = "C5"
    ws.conditional_formatting.add(f"U5:U{last}", CellIsRule(operator="equal", formula=['"Review Required"'],
        fill=PatternFill("solid", fgColor="F5B7B1"), font=Font(bold=True, color="922B21")))
    ws.conditional_formatting.add(f"U5:U{last}", CellIsRule(operator="equal", formula=['"Fully Allocated"'],
        fill=PatternFill("solid", fgColor="D5F5E3"), font=Font(bold=True, color="196F3D")))

    # =====================================================================
    # 5. INVOICE DETAIL (SOA)
    # =====================================================================
    ws = wb.create_sheet("Invoice Detail (SOA)")
    ws.sheet_view.showGridLines = False
    H = ["Customer", "Transaction Type", "Related Sales Order", "Project ID", "P.O. No.", "Document Number",
         "Invoice Date", "Payment Terms (Days)", "Terms Source", "Due Date", "System Due Date",
         "Due Date Variance (Days)", "Age (Days)", "Open Balance", "Current", "0-30", "31-60", "61-90",
         "91-120", "Over 120", "Aging Bucket", "Forecast Month", "Forecast Week", "Data Quality"]
    title(ws, 1, "INVOICE-LEVEL DETAIL / STATEMENT OF ACCOUNT",
          f"Due Date = Invoice Date + Payment Terms (contract terms, from the {PRIOR} master)  |  "
          f"Age = {ASOF_L} less Due Date", span=len(H))
    hdr(ws, 4, H)
    inv_sorted = sorted(INV, key=lambda x: (x["customer"], x["txn_date"] or "9999", x["doc_no"]))
    r = 5
    for x in inv_sorted:
        d = dt.date.fromisoformat(x["txn_date"]) if x["txn_date"] else None
        du = dt.date.fromisoformat(x["due"]) if x["due"] else None
        nd = dt.date.fromisoformat(x["ns_due"]) if x["ns_due"] else None
        bv = [0.0] * 6
        bv[bidx(x["bucket"])] = x["open_bal"]
        vals = [x["customer"], x["txn_type"], x["so"], x["project"], x["po_no"], x["doc_no"],
                d, x["terms"], x["terms_source"], du, nd,
                ((du - nd).days if (du and nd) else None), x["age"], x["open_bal"], *bv,
                x["bucket"], ML.get(x["fc_month"], x["fc_month"]),
                ("W%d" % x["fc_week"]) if x["fc_week"] else "", x["dq"]]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v)
        r += 1
    last = r - 1
    style_body(ws, 5, last, 1, len(H), money_cols=(14, 15, 16, 17, 18, 19, 20), date_cols=(7, 10, 11), banded=False)
    for rr in range(5, last + 1):
        for cc in (1, 9):
            ws.cell(rr, cc).alignment = LEFT
    tr = last + 1
    ws.cell(tr, 14, TOTAL)
    for i in range(6):
        ws.cell(tr, 15 + i, TOTB[i])
    totrow(ws, tr, 1, len(H))
    for c in range(14, 21):
        ws.cell(tr, c).number_format = MONEY
    widths(ws, {"A": 44, "B": 16, "C": 20, "D": 24, "E": 22, "F": 22, "G": 13, "H": 11, "I": 34, "J": 13,
                "K": 14, "L": 13, "M": 11, "N": 16, "O": 15, "P": 14, "Q": 14, "R": 14, "S": 14, "T": 15,
                "U": 13, "V": 14, "W": 12, "X": 30})
    ws.auto_filter.ref = f"A4:{get_column_letter(len(H))}{last}"
    ws.freeze_panes = "B5"
    ws.conditional_formatting.add(f"N5:N{last}", CellIsRule(operator="lessThan", formula=["0"],
        fill=PatternFill("solid", fgColor="FCF3CF"), font=Font(bold=True, color="7D6608")))
    ws.conditional_formatting.add(f"U5:U{last}", CellIsRule(operator="equal", formula=['"Over 120"'],
        fill=PatternFill("solid", fgColor="F5B7B1"), font=Font(bold=True, color="922B21")))
    ws.conditional_formatting.add(f"X5:X{last}", CellIsRule(operator="notEqual", formula=['"OK"'],
        fill=PatternFill("solid", fgColor="FAD7A0")))

    # =====================================================================
    # 6. TERMS MASTER
    # =====================================================================
    ws = wb.create_sheet("Terms Master")
    ws.sheet_view.showGridLines = False
    H = ["S.No", "Customer", "Payment Terms (Days)", "Source of Terms", "Status",
         f"In {PRIOR} File?", f"{PRIOR} Terms", f"{PRIOR} Balance ({CUR})", f"Balance ({CUR})",
         "Documents", "Action Required"]
    title(ws, 1, "PAYMENT TERMS MASTER",
          f"The system does not carry contract payment terms. Terms are carried forward from the {PRIOR} file "
          "by customer name. Confirm every ASSUMED / DERIVED row against the signed contract.", span=len(H))
    hdr(ws, 4, H)
    r = 5
    for i, c in enumerate(sorted(TER, key=lambda x: (TER[x]["status"] == "Confirmed", x)), 1):
        t = TER[c]
        injun = c in PRT
        act = ("None - confirmed from prior master" if t["status"] == "Confirmed"
               else ("Verify against contract - terms derived from system due dates"
                     if t["status"].startswith("Derived")
                     else f"CONFIRM CONTRACT TERMS - {cfg.DEFAULT_TERMS} days assumed"))
        vals = [i, c, t["terms"], t["source"], t["status"], "Yes" if injun else "No",
                PRT.get(c), PRB.get(c), cust[c]["open"], cust[c]["n"], act]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v)
        r += 1
    last = r - 1
    style_body(ws, 5, last, 1, len(H), money_cols=(8, 9))
    for rr in range(5, last + 1):
        for cc in (2, 4, 11):
            ws.cell(rr, cc).alignment = LEFT
    widths(ws, {"A": 6, "B": 48, "C": 12, "D": 38, "E": 26, "F": 14, "G": 11, "H": 17, "I": 17, "J": 11, "K": 46})
    ws.auto_filter.ref = f"A4:K{last}"
    ws.freeze_panes = "C5"
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"Confirmed"'],
        fill=PatternFill("solid", fgColor="D5F5E3"), font=Font(bold=True, color="196F3D")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="containsText",
        formula=['NOT(ISERROR(SEARCH("ASSUMED",E5)))'],
        fill=PatternFill("solid", fgColor="F5B7B1"), font=Font(bold=True, color="922B21")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="containsText",
        formula=['NOT(ISERROR(SEARCH("Derived",E5)))'],
        fill=PatternFill("solid", fgColor="FAD7A0"), font=Font(bold=True, color="9C640C")))

    # =====================================================================
    # 7. EXCEPTIONS
    # =====================================================================
    ws = wb.create_sheet("Exceptions & Mismatches")
    ws.sheet_view.showGridLines = False
    title(ws, 1, "EXCEPTIONS, NAME MISMATCHES & UNAPPLIED CASH",
          "Items requiring manual review before the report is issued", span=12)

    def section(r, text):
        ws.cell(r, 1, text)
        ws.cell(r, 1).font = Font(bold=True, size=11.5, color=WHITE)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=NAVY2)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.cell(r, 1).alignment = CTRN

    r = 4
    section(r, "SECTION A - SIMILAR CUSTOMER NAMES (possible duplicate accounts / payments booked to the wrong master)")
    r += 1
    hdr(ws, r, ["Account A", f"Balance A ({CUR})", "Terms A", "Account B", f"Balance B ({CUR})", "Terms B",
                "Similarity", "Match Type", "Risk", "Recommended Action"])
    r += 1
    a_start = r
    pairs = []
    for g in D["dup_groups"]:
        for i in range(len(g)):
            for j in range(i + 1, len(g)):
                pairs.append((g[i], g[j], 1.0, "Identical after removing legal suffix"))
    for a, b, s in D["near_pairs"]:
        pairs.append((a, b, s, "Near-identical name"))
    for a, b, s, mt in pairs:
        da, db = cust.get(a), cust.get(b)
        if not da or not db:
            continue
        risky = da["neg"] < 0 or db["neg"] < 0
        rk = "HIGH - one side holds unapplied credit" if risky else "MEDIUM"
        act = ("Payment/credit likely booked to the wrong account - reallocate and net off" if risky
               else "Confirm these are separate legal entities; if not, merge in the system")
        for j, v in enumerate([a, da["open"], TER[a]["terms"], b, db["open"], TER[b]["terms"],
                               s, mt, rk, act], 1):
            ws.cell(r, j, v)
        r += 1
    if r - 1 >= a_start:
        style_body(ws, a_start, r - 1, 1, 10, money_cols=(2, 5), pct_cols=(7,))
        for rr in range(a_start, r):
            for cc in (1, 4, 8, 10):
                ws.cell(rr, cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(rr, 7).number_format = "0.0%"

    r += 2
    section(r, "SECTION B - UNAPPLIED CASH, PAYMENTS ON ACCOUNT & CREDIT MEMOS (not matched to an invoice at close)")
    r += 1
    hdr(ws, r, ["Customer", "Transaction Type", "Document Number", "Date", "Due Date", "Age (Days)",
                f"Credit Amount ({CUR})", "Aging Bucket", "Related Sales Order", "Project ID",
                "Customer Net Position", "Recommended Action"])
    r += 1
    b_start = r
    negs = sorted([x for x in INV if x["open_bal"] < 0], key=lambda x: x["open_bal"])
    for x in negs:
        d = cust[x["customer"]]
        act = ("Net off against open invoices for the same customer" if d["open"] > 0
               else "Customer is in net credit - investigate over-collection / duplicate receipt")
        vals = [x["customer"], x["txn_type"], x["doc_no"],
                dt.date.fromisoformat(x["txn_date"]) if x["txn_date"] else None,
                dt.date.fromisoformat(x["due"]) if x["due"] else None,
                x["age"], x["open_bal"], x["bucket"], x["so"], x["project"], d["open"], act]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v)
        r += 1
    if r - 1 >= b_start:
        style_body(ws, b_start, r - 1, 1, 12, money_cols=(7, 11), date_cols=(4, 5), banded=False)
        for rr in range(b_start, r):
            for cc in (1, 12):
                ws.cell(rr, cc).alignment = LEFT
    ws.cell(r, 7, neg)
    totrow(ws, r, 1, 12, label="TOTAL UNAPPLIED CASH / CREDITS")
    ws.cell(r, 7).number_format = MONEY
    r += 3

    section(r, "SECTION C - DATA QUALITY EXCEPTIONS")
    r += 1
    hdr(ws, r, ["Exception", "Count", f"Value ({CUR})", "Impact", "Owner", "Recommended Action"])
    r += 1
    dqs = [
        (f"Payment terms assumed at {cfg.DEFAULT_TERMS} days (no contract on file)",
         sum(1 for v in TER.values() if v["status"].startswith("ASSUMED")),
         sum(cust[c]["open"] for c in TER if TER[c]["status"].startswith("ASSUMED")),
         "Aging bucket may shift once real terms are applied", "Credit Control",
         "Obtain contract terms and update the Terms Master tab"),
        ("Payment terms derived from system due dates",
         sum(1 for v in TER.values() if v["status"].startswith("Derived")),
         sum(cust[c]["open"] for c in TER if TER[c]["status"].startswith("Derived")),
         "Low - derived from posted due dates", "Credit Control", "Spot-check against contract"),
        ("System due date equals invoice date (no terms in system)",
         sum(1 for x in INV if x["ns_due"] and x["txn_date"] and x["ns_due"] == x["txn_date"]),
         sum(x["open_bal"] for x in INV if x["ns_due"] and x["txn_date"] and x["ns_due"] == x["txn_date"]),
         "Root cause of incorrect system aging", "IT / ERP admin",
         "Configure payment terms on the customer master in the ERP"),
        ("Journal entries in the AR sub-ledger",
         sum(1 for x in INV if x["txn_type"] == "Journal"),
         sum(x["open_bal"] for x in INV if x["txn_type"] == "Journal"),
         "Manual adjustments sitting in AR", "Financial Accounting",
         "Review nature and clear or reclassify"),
        ("Documents with no document number",
         sum(1 for x in INV if not x["doc_no"]),
         sum(x["open_bal"] for x in INV if not x["doc_no"]),
         "Cannot be referenced on a statement", "Financial Accounting", "Complete source data"),
        ("Intercompany balances excluded from this report", ic_accounts, ic_total,
         "Excluded to preserve month-on-month comparability", "Group Finance",
         "Reconcile separately via the Intercompany tab"),
    ]
    dq_start = r
    for x in dqs:
        for j, v in enumerate(x, 1):
            ws.cell(r, j, v)
        r += 1
    style_body(ws, dq_start, r - 1, 1, 6, money_cols=(3,))
    for rr in range(dq_start, r):
        for cc in (1, 4, 6):
            ws.cell(rr, cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    widths(ws, {"A": 46, "B": 18, "C": 22, "D": 38, "E": 18, "F": 22, "G": 12, "H": 26, "I": 22, "J": 46,
                "K": 20, "L": 46})

    # =====================================================================
    # 8. INTERCOMPANY (EXCLUDED)
    # =====================================================================
    ws = wb.create_sheet("Intercompany (Excluded)")
    ws.sheet_view.showGridLines = False
    title(ws, 1, "INTERCOMPANY BALANCES - EXCLUDED FROM THIS REPORT",
          f"These {ic_accounts} accounts are excluded from all AR KPIs, aging and forecast, consistent with the "
          f"{PRIOR} presentation. Reconcile through group intercompany.", span=8)
    hdr(ws, 4, ["Intercompany Account", "Documents", f"Open Balance ({CUR})", "% of IC Total",
                "Oldest Document", "Newest Document", "Journals", "Invoices"])
    icg = defaultdict(lambda: dict(n=0, v=0.0, dates=[], j=0, i=0))
    for x in D["ic"]:
        g = icg[x["customer"]]
        g["n"] += 1
        g["v"] += x["open_bal"]
        if x["txn_date"]:
            g["dates"].append(dt.date.fromisoformat(x["txn_date"]))
        if x["txn_type"] == "Journal":
            g["j"] += 1
        if x["txn_type"] == "Invoice":
            g["i"] += 1
    ictot = sum(g["v"] for g in icg.values())
    r = 5
    for c in sorted(icg, key=lambda x: -icg[x]["v"]):
        g = icg[c]
        for j, v in enumerate([c, g["n"], g["v"], g["v"] / ictot if ictot else 0,
                               min(g["dates"]) if g["dates"] else None,
                               max(g["dates"]) if g["dates"] else None, g["j"], g["i"]], 1):
            ws.cell(r, j, v)
        r += 1
    style_body(ws, 5, r - 1, 1, 8, money_cols=(3,), pct_cols=(4,), date_cols=(5, 6))
    for rr in range(5, r):
        ws.cell(rr, 1).alignment = LEFT
    ws.cell(r, 3, ictot)
    ws.cell(r, 2, sum(g["n"] for g in icg.values()))
    ws.cell(r, 4, 1)
    totrow(ws, r, 1, 8, label="TOTAL INTERCOMPANY (EXCLUDED)")
    ws.cell(r, 3).number_format = MONEY
    ws.cell(r, 4).number_format = PCT
    widths(ws, {"A": 46, "B": 13, "C": 22, "D": 13, "E": 17, "F": 17, "G": 12, "H": 12})

    # =====================================================================
    # 9. METHODOLOGY
    # =====================================================================
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    title(ws, 1, "METHODOLOGY, FORMULAS & REFRESH INSTRUCTIONS",
          "How every number in this workbook is produced", span=4)
    hdr(ws, 4, ["#", "Step / Rule", "Formula or Logic", "Why"])
    steps = [
        ("1", "Source data", f"A/R Aging Detail extract as of {ASOF_L}",
         "Single system of record for open items."),
        ("2", "Intercompany exclusion",
         f"Exclude accounts matching the configured IC patterns ({ic_accounts} accounts, "
         f"{CUR} {ic_total:,.2f}, {len(D['ic'])} rows)",
         "Consistent with prior presentation. Including IC would overstate AR and break comparability."),
        ("3", "Payment terms - priority 1", f"Exact customer-name match to the {PRIOR} terms column",
         "The system does not hold contract terms; terms come from the signed contract, maintained manually."),
        ("4", "Payment terms - priority 2", "Normalised name match (case, punctuation, spacing removed)",
         "Catches formatting differences in the customer master."),
        ("5", "Payment terms - priority 3",
         f"Mode of (System Due Date - Invoice Date) per customer, snapped to {cfg.STANDARD_TERMS}",
         "Only for customers not in the prior master. Flagged 'Derived - verify contract'."),
        ("6", "Payment terms - priority 4", f"Default {cfg.DEFAULT_TERMS} days",
         "Last resort. Flagged 'ASSUMED - confirm contract'. Never silently applied."),
        ("7", "Due date", "Due Date = Invoice Date + Payment Terms",
         "The system due date is unusable where terms were never configured (Due Date = Invoice Date rows)."),
        ("8", "Age", f"Age = MAX(0, {ASOF_L} - Due Date)", "Items not yet due show age 0."),
        ("9", "Aging buckets", "Current: age = 0 | 0-30 | 31-60 | 61-90 | 91-120 | Over 120: >120",
         "Bucket boundaries match the historical pack and the collections escalation ladder."),
        ("10", "Bucket reconciliation", "Sum of six buckets - Open AR must equal 0.00 for every customer",
         "Bucket Check column on Customer Summary. Any non-zero is highlighted red."),
        ("11", "Forecast - current items",
         f"Allocated to the calendar month the invoice falls due; within {ML['M1']}, to the week containing "
         "the due date (1-7, 8-14, 15-21, 22-end)",
         "Driven by the real due date rather than judgement."),
        ("12", "Forecast - overdue items",
         f"0-30 to {ML['M1']} week 2 | 31-60 to {ML['M1']} week 4 | 61-90 to {ML['M2']} | "
         f"91-120 to {ML['M3']} | Over 120 to {ML['Beyond']}",
         "Deterministic and fully allocated. Collector may override on the Collection Forecast tab."),
        ("13", "Forecast reconciliation",
         f"{ML['M1']} + {ML['M2']} + {ML['M3']} + Beyond must equal Open AR for every customer",
         "Status column flags any break."),
        ("14", "Risk category",
         "Critical: over-60 share >=75% or oldest >365d | High: >=40% or >180d | Medium: >=15% or >90d | "
         "Low: otherwise | Credit Balance: net credit",
         "Drives the collection priority and the recommended action."),
        ("15", "Collection priority",
         f"P1 over-60 >= {CUR} {cfg.PRIO_P1:,} | P2 >= {cfg.PRIO_P2:,} | P3 > 0 | P4 otherwise",
         "Focuses collector time on value at risk."),
        ("16", "Unapplied cash", "Every document with a negative open balance is listed on Exceptions section B",
         "Receipts and credit memos not matched at close - the main cause of overstated aging."),
        ("17", "Name mismatch review",
         "Names identical after removing legal suffixes, or with >=88% similarity, are paired on Exceptions "
         "section A", "Catches payments posted to a near-duplicate customer master."),
        ("18", "Monthly refresh",
         "Replace the extract, update AS_OF in config.py, re-run run_month_end.py. Carry the Terms Master "
         "tab forward as the new terms source.",
         "Terms are cumulative - each month adds newly confirmed contracts."),
    ]
    r = 5
    for s in steps:
        for j, v in enumerate(s, 1):
            ws.cell(r, j, v)
        r += 1
    style_body(ws, 5, r - 1, 1, 4)
    for rr in range(5, r):
        for cc in (2, 3, 4):
            ws.cell(rr, cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[rr].height = 42
    widths(ws, {"A": 6, "B": 34, "C": 78, "D": 70})

    # =====================================================================
    # 10. RAW DATA
    # =====================================================================
    ws = wb.create_sheet("RAW Data")
    ws.sheet_view.showGridLines = False
    H = ["Customer", "Transaction Type", "Date", "Document Number", "P.O. No.", "System Due Date",
         "System Age", "Open Balance", "Related Sales Order", "Project ID", "Classification", "Source Row"]
    hdr(ws, 1, H)
    r = 2
    for x in INV:
        for j, v in enumerate([x["customer"], x["txn_type"],
                               dt.date.fromisoformat(x["txn_date"]) if x["txn_date"] else None,
                               x["doc_no"], x["po_no"],
                               dt.date.fromisoformat(x["ns_due"]) if x["ns_due"] else None,
                               x["ns_age"], x["open_bal"], x["so"], x["project"], "External", x["src_row"]], 1):
            ws.cell(r, j, v)
        r += 1
    for x in D["ic"]:
        for j, v in enumerate([x["customer"], x["txn_type"],
                               dt.date.fromisoformat(x["txn_date"]) if x["txn_date"] else None,
                               x["doc_no"], x["po_no"],
                               dt.date.fromisoformat(x["ns_due"]) if x["ns_due"] else None,
                               x["ns_age"], x["open_bal"], x["so"], x["project"],
                               "Intercompany - EXCLUDED", x["src_row"]], 1):
            ws.cell(r, j, v)
        r += 1
    style_body(ws, 2, r - 1, 1, 12, money_cols=(8,), date_cols=(3, 6), banded=False)
    for rr in range(2, r):
        ws.cell(rr, 1).alignment = LEFT
    widths(ws, {"A": 44, "B": 16, "C": 13, "D": 22, "E": 22, "F": 16, "G": 12, "H": 16, "I": 20, "J": 24,
                "K": 22, "L": 11})
    ws.auto_filter.ref = f"A1:L{r-1}"
    ws.freeze_panes = "A2"

    for s in wb.worksheets:
        s.sheet_properties.tabColor = NAVY

    wb.save(OUT)
    print("Saved:", OUT)
    print(f"Open AR {TOTAL:,.2f} | buckets {sum(TOTB):,.2f} | forecast {fc_total:,.2f}")
    return OUT


if __name__ == "__main__":
    build()
