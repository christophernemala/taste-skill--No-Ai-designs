"""AR aging core engine.

Reads the NetSuite-style A/R Aging Detail extract and the prior-month terms
master, then:
  1. forward-fills customer names and drops "Total -" rows,
  2. splits intercompany from external (Trap 1),
  3. resolves contract payment terms per the four-level hierarchy (Traps 2-3),
  4. recomputes due date, age and aging bucket,
  5. allocates a deterministic collection forecast over the next 3 months,
  6. builds the exception sets (unapplied cash, near-duplicate names,
     data quality),
and writes a single JSON payload consumed by every downstream builder.

Method reference: sop/AR-SOP.md and the Methodology tab of the workbook.
"""
import datetime as dt
import json
import re
import warnings
from collections import Counter, defaultdict
from difflib import SequenceMatcher

import openpyxl

import sys as _sys
from pathlib import Path as _P
_sys.path.insert(0, str(_P(__file__).resolve().parent.parent))
import config as cfg

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Name normalisation
# ---------------------------------------------------------------------------
LEGAL_SUFFIXES = [
    "PRIVATE JOINT STOCK COMPANY", "PUBLIC JOINT STOCK COMPANY", "JOINT STOCK COMPANY",
    "SOLE PROPRIETORSHIP LLC", "ONE PERSON COMPANY", "LLC SOC", "L L C S O C",
    "FZ LLC", "FZE", "FZC", "DMCC", "DWC", "PJSC", "PSC", "PLC",
    "W L L", "WLL", "L L C", "LLC", "SOC", "SAOC", "SAOG", "SPC",
    "LIMITED", "LTD", "CO", "COMPANY", "EST", "ESTABLISHMENT",
    "GENERAL TRADING", "TRADING", "CONTRACTING", "GROUP", "HOLDING", "HOLDINGS",
]


def norm(name: str) -> str:
    if not name:
        return ""
    s = str(name).upper()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_core(name: str) -> str:
    """Aggressive key: normalised + legal suffixes removed. Flagging ONLY —
    never used to adopt terms or merge accounts."""
    s = norm(name)
    changed = True
    while changed:
        changed = False
        for tok in LEGAL_SUFFIXES:
            if s.endswith(" " + tok):
                s = s[: -(len(tok) + 1)].strip()
                changed = True
            elif s == tok:
                s = ""
                changed = True
    return re.sub(r"\s+", " ", s).strip()


def is_ic(name: str) -> bool:
    if not name:
        return False
    u = str(name).upper()
    return any(p in u for p in cfg.IC_CONTAINS) or any(u.startswith(p) for p in cfg.IC_PREFIXES)


# ---------------------------------------------------------------------------
# 1. Prior-month terms master
# ---------------------------------------------------------------------------
def read_prior_terms():
    """Terms master + prior balances from last month's workbook (SOA sheet:
    customer col 1, terms col 7, open balance col 10, data from row 8).
    Missing file => first run: empty master, everything derived/assumed."""
    if not cfg.PRIOR_FILE.exists():
        return {}, {}
    wb = openpyxl.load_workbook(cfg.PRIOR_FILE, data_only=True)
    ws = wb[cfg.PRIOR_SHEET] if cfg.PRIOR_SHEET in wb.sheetnames else wb.active
    terms, prior_bal = {}, defaultdict(float)
    for r in range(8, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        a = str(a).strip()
        if a.upper().startswith("TOTAL") or a.upper() == "GRAND TOTAL":
            continue
        pt = ws.cell(r, 7).value
        ob = ws.cell(r, 10).value
        if isinstance(pt, (int, float)) and a not in terms:
            terms[a] = int(pt)
        if isinstance(ob, (int, float)):
            prior_bal[a] += ob
    wb.close()
    return terms, dict(prior_bal)


# ---------------------------------------------------------------------------
# 2. Current-month raw detail
# ---------------------------------------------------------------------------
def read_extract():
    wb = openpyxl.load_workbook(cfg.EXTRACT_FILE, data_only=True)
    ws = wb[cfg.EXTRACT_SHEET] if cfg.EXTRACT_SHEET else wb.worksheets[0]
    cust, rows = None, []
    for r in range(8, ws.max_row + 1):
        a = ws.cell(r, 1).value
        tt = ws.cell(r, 2).value
        ob = ws.cell(r, 8).value
        if a is not None:
            a = str(a).strip()
            if a.upper().startswith("TOTAL"):
                continue
            cust = a          # group-header row: remember and move on
            continue
        if tt is None and ob is None:
            continue
        d = ws.cell(r, 3).value
        du = ws.cell(r, 6).value
        rows.append(dict(
            src_row=r,
            customer=cust,
            txn_type=(str(tt).strip() if tt else ""),
            txn_date=(d.date() if isinstance(d, dt.datetime) else d),
            doc_no=(str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value is not None else ""),
            po_no=(str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value is not None else ""),
            ns_due=(du.date() if isinstance(du, dt.datetime) else du),
            ns_age=ws.cell(r, 7).value,
            open_bal=float(ob) if isinstance(ob, (int, float)) else 0.0,
            so=(str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value is not None else ""),
            project=(str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value is not None else ""),
        ))
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# 3. Terms resolution hierarchy
# ---------------------------------------------------------------------------
def derive_terms_from_data(rows):
    """Level 3: mode of (NetSuite Due Date − Invoice Date) per customer,
    ignoring zero-day rows, snapped to the nearest standard term."""
    per = defaultdict(Counter)
    for x in rows:
        if x["txn_type"] != "Invoice":
            continue
        if x["txn_date"] and x["ns_due"]:
            d = (x["ns_due"] - x["txn_date"]).days
            if d > 0:
                per[x["customer"]][d] += 1
    out = {}
    for c, cnt in per.items():
        d = cnt.most_common(1)[0][0]
        out[c] = d if d in cfg.STANDARD_TERMS else min(cfg.STANDARD_TERMS, key=lambda s: abs(s - d))
    return out


def resolve_terms(custs, prior_terms):
    """Levels: 1 exact match, 2 normalised match, 3 derived, 4 default.
    A near-miss name NEVER adopts terms — it is flagged for review instead."""
    by_norm = {}
    for c, t in prior_terms.items():
        by_norm.setdefault(norm(c), t)

    prior_cores = defaultdict(list)
    for c in prior_terms:
        prior_cores[norm_core(c)].append(c)

    resolved, review = {}, []
    for c, derived in custs.items():
        if c in prior_terms:
            resolved[c] = (prior_terms[c], "Prior-month terms master (exact name match)", "Confirmed")
            continue
        n = norm(c)
        if n in by_norm:
            resolved[c] = (by_norm[n], "Prior-month terms master (normalised name match)", "Confirmed")
            continue
        core = norm_core(c)
        cands = list(prior_cores.get(core, []))
        if not cands:
            for pc in prior_terms:
                if SequenceMatcher(None, core, norm_core(pc)).ratio() >= 0.90:
                    cands.append(pc)
        if derived:
            resolved[c] = (derived, "Derived from system (Due Date - Invoice Date)", "Derived - verify contract")
        else:
            resolved[c] = (cfg.DEFAULT_TERMS,
                           f"Default {cfg.DEFAULT_TERMS} days (no contract terms on file)",
                           "ASSUMED - confirm contract")
        if cands:
            review.append((c, cands[:3], resolved[c][0], [prior_terms[x] for x in cands[:3]]))
    return resolved, review


# ---------------------------------------------------------------------------
# 4. Aging and forecast
# ---------------------------------------------------------------------------
def bucket_of(age):
    if age <= 0:
        return "Current"
    if age <= 30:
        return "0-30"
    if age <= 60:
        return "31-60"
    if age <= 90:
        return "61-90"
    if age <= 120:
        return "91-120"
    return "Over 120"


def week_of(d):
    if d.day <= 7:
        return 1
    if d.day <= 14:
        return 2
    if d.day <= 21:
        return 3
    return 4


def forecast_slot(bucket, due):
    """(month_key, week 1-4 or None). month_key in {M1, M2, M3, Beyond}.
    Deterministic: 100% of open AR is allocated; collector overrides live in
    dedicated workbook columns, never in the baseline."""
    if bucket == "Current":
        if due is None:
            return "M1", 4
        k = (due.year, due.month)
        if k == (cfg.M1.year, cfg.M1.month):
            return "M1", week_of(due)
        if k == (cfg.M2.year, cfg.M2.month):
            return "M2", None
        if k == (cfg.M3.year, cfg.M3.month):
            return "M3", None
        if due <= cfg.AS_OF:
            return "M1", week_of(cfg.M1)
        return "Beyond", None
    if bucket == "0-30":
        return "M1", 2
    if bucket == "31-60":
        return "M1", 4
    if bucket == "61-90":
        return "M2", None
    if bucket == "91-120":
        return "M3", None
    return "Beyond", None


# ---------------------------------------------------------------------------
def run():
    cfg.OUT_DIR.mkdir(parents=True, exist_ok=True)
    prior_terms, prior_bal = read_prior_terms()
    raw = read_extract()

    ic_rows = [x for x in raw if is_ic(x["customer"])]
    ext_rows = [x for x in raw if not is_ic(x["customer"])]

    derived = derive_terms_from_data(ext_rows)
    custs = {}
    for x in ext_rows:
        custs.setdefault(x["customer"], derived.get(x["customer"]))
    terms_map, name_review = resolve_terms(custs, prior_terms)

    inv = []
    for x in ext_rows:
        t, src, status = terms_map[x["customer"]]
        due = (x["txn_date"] + dt.timedelta(days=t)) if x["txn_date"] else None
        age = max(0, (cfg.AS_OF - due).days) if due else 0
        b = bucket_of(age)
        mk, wk = forecast_slot(b, due)
        dq = []
        if x["txn_date"] is None:
            dq.append("Missing invoice date")
        if not x["doc_no"]:
            dq.append("Missing document number")
        if status.startswith("ASSUMED"):
            dq.append("Terms assumed")
        if status.startswith("Derived"):
            dq.append("Terms derived")
        rec = dict(x)
        rec.update(
            terms=t, terms_source=src, terms_status=status,
            due=due.isoformat() if due else None,
            age=age, bucket=b, fc_month=mk, fc_week=wk,
            dq=("OK" if not dq else "; ".join(dq)),
            txn_date=x["txn_date"].isoformat() if x["txn_date"] else None,
            ns_due=x["ns_due"].isoformat() if x["ns_due"] else None,
        )
        inv.append(rec)

    unapplied = [r for r in inv if r["open_bal"] < 0]

    cores = defaultdict(list)
    for c in custs:
        cores[norm_core(c)].append(c)
    dup_groups = [sorted(v) for v in cores.values() if len(v) > 1]
    keys = sorted(cores)
    near = []
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            if not keys[i] or not keys[j]:
                continue
            if abs(len(keys[i]) - len(keys[j])) > 8:
                continue
            r = SequenceMatcher(None, keys[i], keys[j]).ratio()
            if 0.88 <= r < 1.0:
                near.append((cores[keys[i]][0], cores[keys[j]][0], round(r, 3)))

    out = dict(
        asof=cfg.AS_OF.isoformat(),
        asof_label=cfg.ASOF_LABEL,
        currency=cfg.CURRENCY,
        company=cfg.COMPANY_NAME,
        prior_label=cfg.PRIOR_LABEL,
        month_labels=cfg.MONTH_LABELS,
        buckets=cfg.BUCKETS,
        invoices=inv,
        ic=[dict(x,
                 txn_date=x["txn_date"].isoformat() if x["txn_date"] else None,
                 ns_due=x["ns_due"].isoformat() if x["ns_due"] else None) for x in ic_rows],
        terms={c: dict(terms=v[0], source=v[1], status=v[2],
                       prior_balance=round(prior_bal.get(c, 0.0), 2))
               for c, v in terms_map.items()},
        prior_terms=prior_terms,
        prior_balances={k: round(v, 2) for k, v in prior_bal.items()},
        name_review=[dict(current=a, prior_candidates=b, applied_terms=c, prior_terms=d)
                     for a, b, c, d in name_review],
        dup_groups=dup_groups,
        near_pairs=near,
        unapplied_count=len(unapplied),
    )
    with open(cfg.ENGINE_JSON, "w") as f:
        json.dump(out, f, default=str)

    # ---- console tie-out ----
    tot = sum(r["open_bal"] for r in inv)
    print(f"External invoices : {len(inv):,}")
    print(f"External customers: {len(custs):,}")
    print(f"Open AR           : {tot:,.2f} {cfg.CURRENCY}")
    print(f"IC excluded       : {sum(x['open_bal'] for x in ic_rows):,.2f} ({len(ic_rows)} rows)")
    bb = Counter()
    for r in inv:
        bb[r["bucket"]] += r["open_bal"]
    for b in cfg.BUCKETS:
        print(f"  {b:<10} {bb[b]:>16,.2f}")
    print(f"  {'CHECK':<10} {sum(bb.values()):>16,.2f}")
    ts = Counter(v[2] for v in terms_map.values())
    print("Terms status:", dict(ts))
    print("Near-name pairs flagged:", len(near), "| dup core groups:", len(dup_groups))
    print(f"Unapplied cash/credit rows: {len(unapplied)}  "
          f"{sum(r['open_bal'] for r in unapplied):,.2f}")
    assert abs(sum(bb.values()) - tot) < 0.01, "Bucket reconciliation failed"
    return out


if __name__ == "__main__":
    run()
